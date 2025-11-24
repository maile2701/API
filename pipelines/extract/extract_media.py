#!/usr/bin/env python
# coding: utf-8

import sys
import subprocess
import os
import time
import requests
import tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageFile
from serpapi import GoogleSearch

ImageFile.LOAD_TRUNCATED_IMAGES = True

# --- Cài đặt package nếu cần ---
def install_package(package):
    subprocess.run([sys.executable, "-m", "pip", "install", package, "--upgrade"])

install_package("google-search-results")
install_package("colabcode")
install_package("openpyxl")
install_package("pillow")

print("✅ Packages đã sẵn sàng!")

# --- Google Search API setup ---
API_KEY = "0e5c803a0b0dd3995df2f264ccc3e89fd686f02fa6833eedb1e66f52db9da7f6"

KEYWORDS = [
    "chùa Thiên Mụ",
    "Kinh đô Huế",
    "Lăng Gia Long",
    "Lăng Minh Mạng",
    "Lăng Tự Đức",
    "Pháp tấn công cửa Thuận An",
    "Festival Huế",
    "Phong trào Cần Vương",
    "Xây dựng cầu Trường Tiền",
    "Vụ mưu khởi nghĩa ở Huế (1916)",
    "Trận Huế Mậu Thân – 1968",
    "Lễ hội điện Hòn Chén",
    "Lễ hội đu tiên",
    "Lễ tế Xã Tắc",
    "Lễ Hội Đền Huyền Trân ở Huế",
    "Pháp bắn phá và chiếm Đà Nẵng 1858",
    "Trận Giải phóng Đà Nẵng 1975",
    "Xây dựng Bà Nà Hills",
    "Xây dựng chùa Linh Ứng (Sơn Trà) Đà Nẵng",
    "Xây dựng cầu quay sông Hàn",
    "Xây dựng cầu Trần Thị Lý",
    "Xây dựng cầu Rồng",
    "Cầu Rồng phun lửa và nước",
    "Xây dựng cầu Thuận Phước",
    "Lễ hội pháo hoa quốc tế tại Đà Nẵng",
    "Lễ hội Carnival -  Bà Nà Hill",
    "Lễ Hội Quán Thế Âm Đà Nẵng"
]

MAX_IMAGES = 10

# --- Hàm crawl ảnh ---
def crawl_images(query, max_images=10):
    print(f"\n🔍 Đang tìm ảnh cho: {query}")
    folder = os.path.join("/Users/thanhmai/etl_pipeline test/data/images", query.replace(" ", "_"))
    os.makedirs(folder, exist_ok=True)

    params = {
        "engine": "google_images",
        "q": query,
        "num": str(max_images),
        "api_key": API_KEY,
    }

    search = GoogleSearch(params)
    results = search.get_dict()

    count = 0
    for i, img in enumerate(results.get("images_results", [])):
        img_url = img.get("original")
        if not img_url:
            continue
        try:
            data = requests.get(img_url, timeout=10).content
            filename = os.path.join(folder, f"{query.replace(' ', '_')}_{i}.jpg")
            with open(filename, "wb") as f:
                f.write(data)
            count += 1
            print(f"✅ {count}/{max_images}: {filename}")
            if count >= max_images:
                break
        except Exception as e:
            print(f"⚠️ Lỗi tải ảnh {i}: {e}")

    print(f" Hoàn tất {count} ảnh cho '{query}'.")

# --- Hàm ghi ảnh vào Excel ---
def save_images_to_excel(root_folder="/Users/thanhmai/etl_pipeline test/data/images", output_path="/Users/thanhmai/etl_pipeline test/data/media_cleaned.xlsx"):
    try:
        if os.path.exists(output_path):
            wb = load_workbook(output_path)
            ws = wb.active
            start_row = ws.max_row + 1
            print(f"📘 Đang cập nhật thêm vào file: {output_path}")
        else:
            raise FileNotFoundError
    except Exception:
        print("⚠️ File Excel không hợp lệ hoặc chưa tồn tại — tạo mới.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách ảnh"
        ws["A1"] = "media_id"
        ws["B1"] = "event_name"
        ws["C1"] = "media"
        ws["D1"] = "media_type"
        start_row = 2

    row = start_row
    existing_names = set()
    if ws.max_row > 1:
        for r in range(2, ws.max_row + 1):
            existing_names.add(ws[f"B{r}"].value)

    temp_files = []
    media_counter = row - 1

    for folder_name in sorted(os.listdir(root_folder)):
        folder_path = os.path.join(root_folder, folder_name)
        if not os.path.isdir(folder_path):
            continue

        print(f"📂 Đang xử lý thư mục: {folder_name}")

        for file in sorted(os.listdir(folder_path)):
            if not file.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')):
                continue

            name = os.path.splitext(file)[0]
            event_name = name.replace("_", " ").strip()
            if event_name in existing_names:
                print(f"⏩ Ảnh {event_name} đã có, bỏ qua.")
                continue

            img_path = os.path.join(folder_path, file)

            try:
                with PILImage.open(img_path) as im:
                    if im.mode in ("RGBA", "LA", "P"):
                        im = im.convert("RGB")
                    im.thumbnail((150, 150))

                    tmpf = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                    tmp_path = tmpf.name
                    tmpf.close()
                    im.save(tmp_path)
                    temp_files.append(tmp_path)

                media_counter += 1
                media_id = f"media_{media_counter:03d}"

                event_lower = event_name.lower()
                media_type = "Du lịch" if any(k in event_lower for k in ["cầu", "lễ hội", "festival"]) else "Lịch sử"

                ws[f"A{row}"] = media_id
                ws[f"B{row}"] = event_name
                img = XLImage(tmp_path)
                ws.add_image(img, f"C{row}")
                ws[f"D{row}"] = media_type

                existing_names.add(event_name)
                row += 1

            except Exception as e:
                print(f"⚠️ Lỗi xử lý ảnh {file}: {e}")
                continue

    try:
        wb.save(output_path)
        wb.close()
        print(f"✅ Đã lưu tất cả ảnh vào file: {output_path}")
    except PermissionError:
        print(f"⚠️ Không thể lưu vì file '{output_path}' đang mở. Hãy đóng Excel rồi chạy lại.")

    for tmp in temp_files:
        try:
            os.remove(tmp)
        except Exception:
            pass

    print("🧹 Đã xóa toàn bộ file tạm thành công.")

# --- Main ---
if __name__ == "__main__":
    for keyword in KEYWORDS:
        crawl_images(keyword, MAX_IMAGES)
        time.sleep(2)

    save_images_to_excel()
    print("\n✅ Toàn bộ quá trình hoàn tất!")
